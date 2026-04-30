#!/usr/bin/env python3
"""Converte contratos.xlsx para data/contratos.js."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from jsonschema import Draft202012Validator
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "scripts" / "schema" / "contratos.schema.json"
DEFAULT_FIELDS = [
    "id",
    "modalidade",
    "numeroModalidade",
    "objeto",
    "processo",
    "contrato",
    "empresa",
    "valor",
    "valorDescricao",
    "dataInicio",
    "dataVencimento",
    "diasPlanilha",
    "status",
    "gestor",
    "fiscal",
    "observacoes",
]
HEADER_ALIASES = {
    "id": "id",
    "modalidade": "modalidade",
    "numero modalidade": "numeroModalidade",
    "numero da modalidade": "numeroModalidade",
    "n modalidade": "numeroModalidade",
    "nº modalidade": "numeroModalidade",
    "objeto": "objeto",
    "processo": "processo",
    "contrato": "contrato",
    "empresa": "empresa",
    "contratada": "empresa",
    "valor": "valor",
    "valor numerico": "valor",
    "valor descrição": "valorDescricao",
    "valor descricao": "valorDescricao",
    "data inicio": "dataInicio",
    "data de inicio": "dataInicio",
    "inicio": "dataInicio",
    "data vencimento": "dataVencimento",
    "data de vencimento": "dataVencimento",
    "vencimento": "dataVencimento",
    "dias planilha": "diasPlanilha",
    "dias": "diasPlanilha",
    "status": "status",
    "situação": "status",
    "situacao": "status",
    "gestor": "gestor",
    "fiscal": "fiscal",
    "observacoes": "observacoes",
    "observações": "observacoes",
}


def main() -> int:
    parser = argparse.ArgumentParser(description="Converte planilha de contratos para data/contratos.js.")
    parser.add_argument("input", nargs="?", default="contratos.xlsx", help="Caminho da planilha .xlsx.")
    parser.add_argument("--sheet", default="CONTRATOS", help="Nome da aba da planilha.")
    parser.add_argument("--output", default="data/contratos.js", help="Arquivo JS de saída.")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    dataset = build_dataset(input_path, args.sheet)
    validate_dataset(dataset)
    write_js(output_path, dataset)
    print(f"{dataset['recordCount']} registro(s) gravado(s) em {output_path}.")
    return 0


def build_dataset(input_path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise SystemExit(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {available}")

    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise SystemExit("A planilha está vazia.") from exc

    header_map = build_header_map(header_row)
    records = []
    for row in rows:
        if is_blank_row(row):
            continue
        record = normalize_row(row, header_map, len(records) + 1)
        records.append(record)

    records.sort(key=lambda item: int(item["id"]))
    return {
        "source": input_path.name,
        "sheet": sheet_name,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "recordCount": len(records),
        "records": records,
    }


def build_header_map(header_row: tuple[Any, ...]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    for index, value in enumerate(header_row):
        key = normalize_header(value)
        field = HEADER_ALIASES.get(key)
        if field:
            mapped[index] = field
    return mapped


def normalize_row(row: tuple[Any, ...], header_map: dict[int, str], fallback_id: int) -> dict[str, Any]:
    raw = {field: None for field in DEFAULT_FIELDS}
    for index, field in header_map.items():
        raw[field] = normalize_cell(row[index] if index < len(row) else None)

    raw["id"] = parse_int(raw["id"]) or fallback_id
    raw["valor"] = parse_money(raw["valor"])
    raw["diasPlanilha"] = parse_number(raw["diasPlanilha"])
    raw["dataInicio"] = parse_date(raw["dataInicio"])
    raw["dataVencimento"] = parse_date(raw["dataVencimento"])
    return raw


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value).strip()
        return text or None
    return value


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    return int(number) if number is not None else None


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None


def parse_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = re.sub(r"[^\d,.-]", "", str(value)).replace(".", "").replace(",", ".")
    try:
        return round(float(Decimal(cleaned)), 2)
    except (InvalidOperation, ValueError):
        return None


def parse_date(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def validate_dataset(dataset: dict[str, Any]) -> None:
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    validator = Draft202012Validator(schema)
    errors = sorted(validator.iter_errors(dataset), key=lambda error: list(error.path))
    if errors:
        for error in errors:
            location = "/" + "/".join(str(part) for part in error.path)
            print(f"Erro de schema em {location}: {error.message}")
        raise SystemExit(1)


def write_js(output_path: Path, dataset: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    content = (
        "/* Arquivo gerado por scripts/xlsx_to_js.py. Não edite manualmente. */\n"
        "window.CONTRATOS_DATA = "
        f"{json.dumps(dataset, ensure_ascii=False, indent=2)};\n"
    )
    output_path.write_text(content, encoding="utf-8", newline="\n")


def normalize_header(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip().lower()
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(normalize_cell(value) is None for value in row)


if __name__ == "__main__":
    raise SystemExit(main())
